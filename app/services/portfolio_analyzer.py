import io
import json
import openpyxl
from datetime import datetime, date

DEMO_USER_AUTH0_ID = "demo|pulseboard"
DEMO_COMPANY_NAME = "NimbusTech Inc. (Sample Model)"


# ---------------------------------------------------------------------------
# Excel parsing
# ---------------------------------------------------------------------------

_COL_MAP = {
    "company name": "company_name", "company": "company_name", "name": "company_name",
    "ticker": "ticker", "symbol": "ticker",
    "sector": "sector", "industry": "sector",
    "investment type": "investment_type", "type": "investment_type", "instrument": "investment_type",
    "shares": "shares", "units": "shares", "quantity": "shares",
    "cost per share": "cost_per_share", "cost/share": "cost_per_share",
    "price paid": "cost_per_share", "purchase price": "cost_per_share",
    "investment date": "investment_date", "date": "investment_date", "purchase date": "investment_date",
    "private valuation ($)": "private_valuation_usd",
    "private valuation": "private_valuation_usd",
    "valuation": "private_valuation_usd",
    "current valuation": "private_valuation_usd",
    "company valuation": "private_valuation_usd",
    "notes": "notes", "note": "notes", "comments": "notes",
}


def parse_excel(file_bytes: bytes) -> tuple[list[dict], list[str]]:
    """
    Parse uploaded Excel bytes into holding dicts.
    Returns (holdings_list, error_list).
    Expects a header row; remaining rows are holdings.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
    col_indices: dict[str, int] = {}
    for i, cell in enumerate(header_row):
        key = str(cell or "").strip().lower()
        canonical = _COL_MAP.get(key)
        if canonical and canonical not in col_indices:
            col_indices[canonical] = i

    holdings: list[dict] = []
    errors: list[str] = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue

        def get(key):
            idx = col_indices.get(key)
            return row[idx] if idx is not None and idx < len(row) else None

        company_name = str(get("company_name") or "").strip()
        if not company_name:
            continue

        try:
            shares_raw = get("shares")
            shares = float(shares_raw) if shares_raw is not None else None

            cost_raw = get("cost_per_share")
            cost_per_share = float(cost_raw) if cost_raw is not None else None

            inv_date = get("investment_date")
            if isinstance(inv_date, datetime):
                investment_date = inv_date.date()
            elif isinstance(inv_date, date):
                investment_date = inv_date
            elif isinstance(inv_date, str) and inv_date.strip():
                try:
                    investment_date = date.fromisoformat(inv_date.strip())
                except ValueError:
                    investment_date = None
            else:
                investment_date = None

            priv_raw = get("private_valuation_usd")
            if priv_raw is not None and str(priv_raw).strip() not in ("", "0"):
                private_valuation_usd = int(float(str(priv_raw).replace(",", "")))
            else:
                private_valuation_usd = None

            holdings.append({
                "company_name": company_name,
                "ticker": str(get("ticker") or "").strip().upper() or None,
                "sector": str(get("sector") or "").strip() or None,
                "investment_type": str(get("investment_type") or "").strip() or None,
                "shares": shares,
                "cost_per_share": cost_per_share,
                "investment_date": investment_date,
                "private_valuation_usd": private_valuation_usd,
                "notes": str(get("notes") or "").strip() or None,
            })
        except (ValueError, TypeError) as exc:
            errors.append(f"Row {row_num} ({company_name}): {exc}")

    return holdings, errors


# ---------------------------------------------------------------------------
# Market data via yfinance
# ---------------------------------------------------------------------------

def fetch_market_data(ticker: str) -> dict:
    """Fetch current price and market cap from Yahoo Finance. Returns {} on failure."""
    try:
        import yfinance as yf
        obj = yf.Ticker(ticker)
        fi = obj.fast_info
        price = fi.last_price
        cap = getattr(fi, "market_cap", None)
        return {
            "last_price": float(price) if price else None,
            "market_cap": int(cap) if cap and cap > 0 else None,
        }
    except Exception:
        return {}


# ---------------------------------------------------------------------------
# Analysis
# ---------------------------------------------------------------------------

def build_holding_analysis(holding) -> dict:
    """Compute per-holding metrics using cached price data."""
    shares = holding.shares or 0
    total_cost = shares * (holding.cost_per_share or 0)
    is_public = bool(holding.ticker)
    current_price = holding.last_price

    if is_public and current_price:
        current_value = shares * current_price
        pnl = current_value - total_cost
        pnl_pct = (pnl / total_cost * 100) if total_cost else None
    else:
        current_value = None
        pnl = None
        pnl_pct = None

    return {
        "holding": holding,
        "total_cost": total_cost,
        "is_public": is_public,
        "current_price": current_price,
        "current_value": current_value,
        "pnl": pnl,
        "pnl_pct": pnl_pct,
    }


# ---------------------------------------------------------------------------
# Valuation model parsing (360 Huntington Fund template format)
# ---------------------------------------------------------------------------

def parse_valuation_model(file_bytes: bytes) -> dict | None:
    """
    Parse a 360 Huntington Fund-style valuation model Excel.
    Financial values in IS Summary are in millions ($M) despite $K labeling.
    Returns extracted dict, or None if the format isn't recognized.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheets = {s.lower(): s for s in wb.sheetnames}

    if "is summary" not in sheets and "dcf" not in sheets:
        return None

    result: dict = {}
    today_year = date.today().year

    # ── Assumptions ──
    if "assumptions" in sheets:
        ws = wb[sheets["assumptions"]]
        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            label = str(row[0]).strip().lower()
            val = row[1] if len(row) > 1 else None
            if not isinstance(val, (int, float)):
                continue
            if "current share price" in label:
                result["model_price"] = round(float(val), 2)
            elif "risk-free rate" in label:
                result["risk_free_rate"] = round(float(val), 4)
            elif "equity beta" in label:
                result["beta"] = round(float(val), 4)
            elif "capm cost of equity" in label:
                result["cost_of_equity"] = round(float(val), 4)

    # ── DCF ──
    if "dcf" in sheets:
        ws = wb[sheets["dcf"]]
        for row in ws.iter_rows(values_only=True):
            if not row or len(row) < 2:
                continue
            label = str(row[1] or "").strip().lower()
            val_c = row[2] if len(row) > 2 else None
            val_d = row[3] if len(row) > 3 else None
            if label == "wacc":
                v = val_d if isinstance(val_d, (int, float)) else (val_c if isinstance(val_c, (int, float)) else None)
                if v: result["wacc"] = round(float(v), 4)
            elif "perpetuity growth rate" in label and isinstance(val_c, (int, float)):
                result["terminal_growth_rate"] = round(float(val_c), 4)
            elif "equity value (present value)" in label and isinstance(val_d, (int, float)):
                result["dcf_per_share"] = round(float(val_d), 2)

    # ── IS Summary ──
    if "is summary" in sheets:
        ws = wb[sheets["is summary"]]
        fiscal_years: list[int] = []

        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            label = str(row[0] or "").strip()

            if label == "Fiscal Year:":
                for v in row[1:]:
                    if isinstance(v, datetime):
                        fiscal_years.append(v.year)
                    elif isinstance(v, date):
                        fiscal_years.append(v.year)
                continue

            if not fiscal_years:
                continue

            n = len(fiscal_years)
            raw = [row[i + 1] if i + 1 < len(row) else None for i in range(n)]
            floats = [round(float(v), 4) if isinstance(v, (int, float)) else None for v in raw]

            if label == "Net Sales":
                result["revenue"] = floats
            elif label == "Operating Income (EBIT)":
                result["ebit"] = floats
            elif label == "Net Income":
                result["net_income"] = floats
            elif label == "Gross Margin %":
                result["gross_margin"] = floats
            elif label == "Operating Margin %":
                result["ebit_margin"] = floats

        if fiscal_years:
            result["fiscal_years"] = fiscal_years
            result["is_projected"] = [yr >= today_year for yr in fiscal_years]

    # ── Summary ──
    if "summary" in sheets:
        ws = wb[sheets["summary"]]
        for row in ws.iter_rows(values_only=True):
            if not row or len(row) < 2:
                continue
            label = str(row[1] or "").strip().lower()
            if "target price (12-month)" in label:
                if len(row) > 4 and isinstance(row[4], (int, float)):
                    result["target_low"] = round(float(row[4]), 2)
                if len(row) > 5 and isinstance(row[5], (int, float)):
                    result["target_median"] = round(float(row[5]), 2)
                if len(row) > 6 and isinstance(row[6], (int, float)):
                    result["target_high"] = round(float(row[6]), 2)
            elif label == "current price" and len(row) > 5 and isinstance(row[5], (int, float)):
                result["model_current_price"] = round(float(row[5]), 2)
            elif "upside from current price" in label and len(row) > 5 and isinstance(row[5], (int, float)):
                result["model_upside"] = round(float(row[5]), 4)

    return result if len(result) > 1 else None


# ---------------------------------------------------------------------------
# Guest-viewable demo model
# ---------------------------------------------------------------------------

def seed_demo_valuation():
    """Seed a permanent sample valuation model, owned by a system demo user,
    so guests can see the Valuation Model Analysis feature without an account."""
    from ..extensions import db
    from ..models.models import User, ValuationAnalysis

    demo_user = User.query.filter_by(auth0_id=DEMO_USER_AUTH0_ID).first()
    if not demo_user:
        demo_user = User(
            auth0_id=DEMO_USER_AUTH0_ID,
            email="demo@pulseboard.io",
            name="PulseBoard Demo",
        )
        db.session.add(demo_user)
        db.session.flush()

    existing = ValuationAnalysis.query.filter_by(
        user_id=demo_user.id, company_name=DEMO_COMPANY_NAME,
    ).first()
    if existing:
        return

    model_data = {
        "model_price": 42.50,
        "risk_free_rate": 0.0425,
        "beta": 1.15,
        "cost_of_equity": 0.0950,
        "wacc": 0.0875,
        "terminal_growth_rate": 0.025,
        "dcf_per_share": 47.80,
        "fiscal_years": [2023, 2024, 2025, 2026, 2027],
        "is_projected": [False, False, True, True, True],
        "revenue": [820.4, 968.1, 1120.6, 1290.7, 1475.9],
        "ebit": [102.5, 145.2, 179.3, 219.4, 265.7],
        "net_income": [68.3, 98.7, 124.5, 153.2, 187.6],
        "gross_margin": [0.712, 0.724, 0.731, 0.738, 0.744],
        "ebit_margin": [0.125, 0.150, 0.160, 0.170, 0.180],
        "target_low": 38.00,
        "target_median": 45.50,
        "target_high": 53.00,
        "model_current_price": 42.50,
        "model_upside": 0.0706,
    }

    analysis = ValuationAnalysis(
        user_id=demo_user.id,
        company_name=DEMO_COMPANY_NAME,
        ticker=None,
        data_json=json.dumps(model_data),
    )
    db.session.add(analysis)
    db.session.commit()


def build_portfolio_summary(analyses: list) -> dict:
    """Compute portfolio-level aggregates from per-holding analyses."""
    total_cost = sum(a["total_cost"] for a in analyses)
    valued = [a for a in analyses if a["current_value"] is not None]
    total_current = sum(a["current_value"] for a in valued)
    total_pnl = sum(a["pnl"] for a in valued if a["pnl"] is not None)
    pnl_cost_base = sum(a["total_cost"] for a in valued)

    sectors: dict[str, float] = {}
    for a in analyses:
        sector = a["holding"].sector or "Unknown"
        sectors[sector] = sectors.get(sector, 0) + a["total_cost"]

    return {
        "total_cost": total_cost,
        "total_current_value": total_current,
        "total_pnl": total_pnl if valued else None,
        "pnl_pct": (total_pnl / pnl_cost_base * 100) if pnl_cost_base else None,
        "num_holdings": len(analyses),
        "num_public": len([a for a in analyses if a["is_public"]]),
        "num_private": len([a for a in analyses if not a["is_public"]]),
        "sectors": dict(sorted(sectors.items(), key=lambda x: x[1], reverse=True)),
    }
